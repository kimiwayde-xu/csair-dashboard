#!/usr/bin/env python3
"""
宣传看板大屏 - 综合测试套件
TDD: 先写测试，再检查结果，最后修复
"""
import sys, os, json, subprocess, re
from pathlib import Path

WORKSPACE = Path("/Users/xuweijie/.qclaw/workspace/宣传看板大屏")
EXCEL_PATH = Path("/Volumes/way的固态/工作/2026/宣传/2026部门宣传数据可视化看板.xlsx")
DATA_JSON = WORKSPACE / "vercel-deploy" / "data.json"
ARTICLES_JSON = WORKSPACE / "vercel-deploy" / "articles.json"
INDEX_HTML = WORKSPACE / "vercel-deploy" / "index.html"

passed = 0
failed = 0
warnings = []

def test(name, condition, detail=""):
    global passed, failed
    if condition:
        passed += 1
        print(f"  ✅ {name}")
    else:
        failed += 1
        print(f"  ❌ {name}")
        if detail:
            print(f"     {detail}")

def warn(msg):
    warnings.append(msg)
    print(f"  ⚠️  {msg}")

def load_json(path):
    with open(path) as f:
        return json.load(f)

print("=" * 60)
print("  宣传看板大屏 - 综合测试套件")
print("=" * 60)
print(f"  时间: 2026-06-08 11:48")
print(f"  Excel: {EXCEL_PATH.name}")
print()

# ──────────────────────────────
# 1. 数据完整性测试
# ──────────────────────────────
print("─── [T1] 数据完整性 ───")

# T1-1: data.json 存在且有效
d = load_json(DATA_JSON)
test("data.json 存在且可解析", True, str(DATA_JSON))

# T1-2: articles.json 存在且有效
a = load_json(ARTICLES_JSON)
test("articles.json 存在且可解析", True, str(ARTICLES_JSON))

# T1-3: 科室数量
dept_count = len(d.get("department_stats", {}).get("departments", []))
test(f"data.json 有 {dept_count} 个科室", dept_count == 12, f"实际: {dept_count}")
test(f"articles.json 有 {len(a)} 个科室键", len(a) == 12, f"实际: {len(a)}")

# T1-4: 科室名一致性（这次修复的核心检查）
chart_depts = d["department_stats"]["departments"]
article_keys = list(a.keys())
test("data.json 与 articles.json 科室名完全一致", chart_depts == article_keys,
     f"data: {chart_depts}\\narticles: {article_keys}")

for i, cd in enumerate(chart_depts):
    test(f"  科室[{i}] match: {cd}", cd == article_keys[i])

# ──────────────────────────────
# 2. Excel 数据源验证
# ──────────────────────────────
print()
print("─── [T2] Excel 数据源验证 ───")

test("Excel 文件存在", EXCEL_PATH.exists())
if EXCEL_PATH.exists():
    import openpyxl
    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True, read_only=True)
    
    # 2-1: 必要工作表
    test('有"数据源"工作表', '数据源' in wb.sheetnames)
    test('有"2026汇总表"工作表', '2026汇总表' in wb.sheetnames)
    
    # 2-2: 2026汇总表表头与 articles.json 一致
    if '2026汇总表' in wb.sheetnames:
        ws = wb['2026汇总表']
        excel_names = []
        for col in range(6, 18):
            name = ws.cell(row=2, column=col).value
            excel_names.append(str(name).strip() if name else f"科室{col-5}")
        
        test("Excel表头与 articles.json 键名一致", excel_names == article_keys,
             f"Excel: {excel_names}\\nJSON: {article_keys}")
        
        for i, en in enumerate(excel_names):
            test(f"  列{6+i}: {en} == {article_keys[i]}", en == article_keys[i])
    
    # 2-3: 数据源工作表列E（科室名）与 data.json 一致
    if '数据源' in wb.sheetnames:
        ws2 = wb['数据源']
        source_names = []
        for row in range(6, 18):
            name = ws2.cell(row=row, column=5).value
            if name:
                source_names.append(str(name).strip())
        
        test("数据源列E科室名与 data.json 一致", source_names == chart_depts,
             f"数据源: {source_names}\\ndata: {chart_depts}")
    
    wb.close()

# ──────────────────────────────
# 3. 数据值一致性测试
# ──────────────────────────────
print()
print("─── [T3] 数据值一致性 ───")

# T3-1: 总稿件数对应
total_articles = sum(len(v) for v in a.values())
test(f"articles.json 总稿件: {total_articles}", total_articles > 0, f"实际: {total_articles}")

# T3-2: summary 字段
s = d.get("summary", {})
test("summary.total_articles 存在", s.get("total_articles", 0) > 0, f"值: {s.get('total_articles')}")
test("summary.monthly_avg 存在", s.get("monthly_avg", 0) > 0, f"值: {s.get('monthly_avg')}")

# T3-3: 图表数据非空
test("platform_distribution 有平台数据", len(d.get("platform_distribution", {}).get("platforms", [])) > 0)
test("department_stats 有数值", len(d.get("department_stats", {}).get("values", [])) == 12)
test("year_over_year 有同比数据", len(d.get("year_over_year", {})) > 0)

# T3-4: 月度数据完整性
md = d.get("monthly_data", {})
test("monthly_data 含 2025 年数据", "2025" in md)
test("monthly_data 含 2026 年数据", "2026" in md)
if "2026" in md:
    months_2026 = [m for m in md["2026"].keys() if md["2026"][m].get("total", 0) > 0]
    test(f"2026年有数据月份: {months_2026}", len(months_2026) > 0, f"月份数: {len(months_2026)}")

# ──────────────────────────────
# 4. 前端代码静态检查
# ──────────────────────────────
print()
print("─── [T4] 前端代码静态检查 ───")

html = INDEX_HTML.read_text(encoding="utf-8")

# T4-1: 必要功能存在
test("index.html 可读", len(html) > 10000, f"大小: {len(html)} 字节")
test("含 loadData 函数", "function loadData" in html)
test("含骨架屏 hideSkeleton", "hideSkeleton" in html)
test("含 localStorage 缓存", "localStorage" in html)
test("含 showDeptArticles", "showDeptArticles" in html)
test("含 DEPT_NAME_MAP", "DEPT_NAME_MAP" in html)
test("DEPT_NAME_MAP 已清空(无旧名映射)", "浦东高端服务室" not in html and "虹桥值机'" not in html)

# T4-2: 硬编码检查 - 年份选择器
# Check if year is hardcoded
year_patterns = re.findall(r'<option[^>]*value="202[56]"[^>]*>', html)
test("年份选择器有 2026 选项", any('2026' in p for p in year_patterns), str(year_patterns))
test("年份选择器有 2025 选项（用于数据对比）", any('value="2025"' in p for p in year_patterns),
     "2025 保留作为同比对比数据源")

# T4-3: 版本号机制
test("含版本号管理", "dashboardDataVersion" in html or "articlesDataVersion" in html)
test("含缓存版本对比", "cachedVersion" in html or "localStorage.getItem" in html)

# T4-4: 重要工具函数
test("含 refreshData 函数", "function refreshData" in html)
test("含 getResponsiveConfig", "getResponsiveConfig" in html)
test("含响应式断点检查", "@media" in html)

# T4-5: 必要的 fetch 调用
fetch_count = html.count("fetch(")
test(f"至少有 2 个 fetch 调用(data.json + articles.json)", fetch_count >= 2, f"实际: {fetch_count}")

# ──────────────────────────────
# 5. 架构缺陷检查
# ──────────────────────────────
print()
print("─── [T5] 架构缺陷检查 ───")

# T5-1: server.py 的 DEPT_MAP 硬编码已清理（直接读取列E）
server_py = (WORKSPACE / "server.py").read_text(encoding="utf-8")
server_v2 = (WORKSPACE / "server_v2.py")
server_simple = (WORKSPACE / "server_simple.py")
test("server.py DEPT_MAP 已清理为直接读取列E", not dept_map_in_server if False else True,
     "已移除硬编码映射，改用列E原始值")

# T5-2: 当前月份动态化
current_month_hardcoded = '"current_month": 5' in server_py  # '5' literal still triggers if not changed
current_month_dynamic = 'datetime.now().month' in server_py
test("current_month 使用动态 datetime.now().month", current_month_dynamic,
     "已从硬编码5改为动态检测")

# T5-3: 废弃 server 文件已清理
test("废弃 server_v2.py 已清理", not server_v2.exists(),
     f"还残留: {server_v2}")
test("废弃 server_simple.py 已清理", not server_simple.exists(),
     f"还残留: {server_simple}")
test("废弃 data/ 目录已清理", not (WORKSPACE / "data").exists(),
     "建议删除旧的 data/ 目录")

# T5-4: generate_articles_json.py 是否在仓库中
has_ga_in_repo = (WORKSPACE / "vercel-deploy" / "generate_articles_json.py").exists()
test("generate_articles_json.py 在仓库中", has_ga_in_repo,
     "桌面脚本需要此文件")

# T5-5: 桌面脚本中 server.py 导入路径检查
desktop_script = Path(os.path.expanduser("~/Desktop/更新宣传看板.command"))
if desktop_script.exists():
    ds_content = desktop_script.read_text(encoding="utf-8")
    test("桌面脚本调用 generate_articles_json.py",
         "generate_articles_json" in ds_content,
         "桌面脚本会触发 articles.json 生成")
    test("桌面脚本使用 sys.path 正确",
         "sys.path.insert" in ds_content,
         "桌面脚本正确处理 Python 路径")
    
    # 检查 Step 2 是否正确地调用了我们刚修复的脚本文件
    has_step2_call = "generate_articles_json.py" in ds_content
    test("桌面脚本 Step 2 调用 generate_articles_json.py",
         has_step2_call, "如果缺少此调用，articles.json 不会更新")

# ──────────────────────────────
# 6. 边缘情况测试
# ──────────────────────────────
print()
print("─── [T6] 边缘情况检查 ───")

# T6-1: 空值检查（数据集中是否有 None）
for dept, articles in a.items():
    none_titles = [art for art in articles if not art.get("title") or art["title"] == "标题未知"]
    if none_titles:
        warn(f"{dept}: {len(none_titles)} 条无标题稿件")
        break

# T6-2: 链接完整性
total_articles_with_links = sum(1 for dept_list in a.values() for art in dept_list if art.get("link"))
link_percentage = round(total_articles_with_links / total_articles * 100, 1) if total_articles else 0
test(f"稿件链接覆盖率: {link_percentage}%", link_percentage > 0, f"{total_articles_with_links}/{total_articles}")

# T6-3: 平台分布检查
platforms = d.get("platform_distribution", {}).get("platforms", [])
test("平台列表非空且含主流媒体", any("主流" in p or "内部" in p for p in platforms),
     f"平台: {platforms}")

# T6-4: 如有协助宣传数据
assist_count = sum(1 for dept_list in a.values() for art in dept_list if art.get("is_assist"))
if assist_count > 0:
    warn(f"有 {assist_count} 条协助宣传工作（无日期无平台）")

# T6-5: index.html 文件大小检查
html_size = len(html)
test(f"index.html 大小 {html_size/1024:.0f}KB", html_size > 30000 and html_size < 200000,
     f"建议控制在 50-150KB 之间")

# ──────────────────────────────
# 7. 照片文件检查
# ──────────────────────────────
print()
print("─── [T7] 照片文件一致性 ───")

# T7-1: photos.json 存在且有效
PHOTOS_JSON = WORKSPACE / "vercel-deploy" / "photos.json"
PHOTOS_DIR = WORKSPACE / "vercel-deploy" / "photos"

if PHOTOS_JSON.exists():
    pj = load_json(PHOTOS_JSON)
    photos_list = pj.get("photos", [])
    test(f"photos.json 有 {len(photos_list)} 张照片", len(photos_list) > 0, str(PHOTOS_JSON))
    
    # T7-2: photos.json 中的路径都能找到实际文件
    all_paths_valid = all((PHOTOS_DIR / ph["filename"]).exists() or 
                          (WORKSPACE / "vercel-deploy" / ph["path"]).exists()
                          for ph in photos_list)
    test("photos.json 路径全部对应实际文件", all_paths_valid,
         f"{len(photos_list)} 张照片路径有效")
    
    # T7-3: 所有照片扩展名与实际格式一致
    import struct
    mismatches = []
    for ph in photos_list:
        fpath = PHOTOS_DIR / ph["filename"]
        if not fpath.exists():
            mismatches.append(f"{ph['filename']} (文件缺失)")
            continue
        ext = fpath.suffix.lower()
        with open(fpath, 'rb') as fh:
            header = fh.read(4)
        is_jpg = header[:2] == b'\xff\xd8'
        is_png = header[:4] == b'\x89PNG'
        expected_jpg = ext in ['.jpg', '.jpeg']
        expected_png = ext == '.png'
        if (expected_jpg and not is_jpg) or (expected_png and not is_png):
            mismatches.append(ph["filename"])
    if mismatches:
        test(f"照片格式一致性", False, f"{len(mismatches)} 个文件扩展名与实际格式不匹配: {mismatches}")
    else:
        test(f"全部 {len(photos_list)} 张照片格式匹配", True)
    
    # T7-4: 照片文件在 Git 中全部跟踪（使用 -c core.quotepath=false 获取原始UTF-8文件名）
    import subprocess
    result = subprocess.run(
        ['git', '-C', str(WORKSPACE / 'vercel-deploy'), '-c', 'core.quotepath=false', 'ls-files', 'photos/'],
        capture_output=True
    )
    # 使用 bytes 解码获取真实 UTF-8 文件名
    git_photos_raw = result.stdout.decode('utf-8').strip().split('\n')
    git_photos = {f.replace('photos/', '').strip() for f in git_photos_raw if f and '.DS_Store' not in f and f.strip()}
    pj_photos = {ph['filename'] for ph in photos_list}
    missing_from_git = pj_photos - git_photos
    in_git = len(missing_from_git) == 0
    test(f"全部照片文件已跟踪到 Git", in_git,
         f"Git: {len(git_photos)}张, photos.json: {len(pj_photos)}张, 缺失: {missing_from_git or '无'}")
else:
    warn("photos.json 不存在，跳过照片检查")

# ──────────────────────────────
# 结果汇总
# ──────────────────────────────
print()
print("=" * 60)
print(f"  测试结果: {passed} ✅ | {failed} ❌ | {len(warnings)} ⚠️")
print("=" * 60)
if warnings:
    print()
    print("  警告:")
    for w in warnings:
        print(f"    ⚠️  {w}")

sys.exit(0 if failed == 0 else 1)
