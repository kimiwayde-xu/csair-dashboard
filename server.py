#!/usr/bin/env python3
"""
宣传看板数据更新服务
启动后访问 http://127.0.0.1:8765 即可使用 Web 界面上传 Excel 并自动部署
"""

import json
import os
import subprocess
import sys
import re
from http.server import HTTPServer, SimpleHTTPRequestHandler
from pathlib import Path
import cgi
from datetime import datetime

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

PORT = 8765
WORKSPACE = Path(__file__).parent
DEPLOY_DIR = WORKSPACE / "vercel-deploy"

# 科室名称读取策略：
# - 数据源工作表列E已是全称（如'虹桥值机室'），直接使用
# - 如需数据源与2026汇总表一致，两表都改即可
# - 废弃说明：旧版DEPT_MAP硬编码短名→全称映射，与列E全称重复，已于2026-06-08移除

class UploadHandler(SimpleHTTPRequestHandler):
    """处理文件上传和数据解析"""

    def do_GET(self):
        if self.path == '/':
            self.send_html()
        elif self.path == '/status':
            self.send_status()
        else:
            super().do_GET()

    def do_POST(self):
        if self.path == '/upload':
            self.handle_upload()
        elif self.path == '/deploy':
            self.handle_deploy()
        else:
            self.send_error(404)

    def send_html(self):
        html = get_html_template()
        self.send_response(200)
        self.send_header('Content-Type', 'text/html; charset=utf-8')
        self.send_header('Content-Length', len(html.encode()))
        self.end_headers()
        self.wfile.write(html.encode())

    def send_status(self):
        status = {
            'status': 'ok',
            'has_openpyxl': HAS_OPENPYXL,
            'workspace': str(WORKSPACE),
        }
        self.send_json(status)

    def send_json(self, data):
        body = json.dumps(data, ensure_ascii=False).encode()
        self.send_response(200)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Content-Length', len(body))
        self.end_headers()
        self.wfile.write(body)

    def handle_upload(self):
        content_type, _ = cgi.parse_header(self.headers['Content-Type'])
        if content_type != 'multipart/form-data':
            self.send_error(400, 'Expected multipart/form-data')
            return

        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={'REQUEST_METHOD': 'POST'}
        )

        if 'file' not in form:
            self.send_error(400, 'No file uploaded')
            return

        file_item = form['file']
        if not file_item.filename:
            self.send_error(400, 'Empty file')
            return

        upload_path = WORKSPACE / 'uploaded.xlsx'
        with open(upload_path, 'wb') as f:
            f.write(file_item.file.read())

        try:
            data = parse_excel(upload_path)

            json_path = WORKSPACE / 'data' / 'dashboard_data.json'
            json_path.parent.mkdir(exist_ok=True)
            with open(json_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            self.send_json({
                'success': True,
                'message': f'解析成功：{file_item.filename}',
                'summary': data.get('summary', {}),
            })
        except Exception as e:
            import traceback
            self.send_json({
                'success': False,
                'error': f'{str(e)}\n{traceback.format_exc()}'
            })

    def handle_deploy(self):
        try:
            # ===== 新架构(2026-06-08): 只更新data.json+articles.json,不修改index.html =====
            json_path = WORKSPACE / 'data' / 'dashboard_data.json'
            if not json_path.exists():
                self.send_json({'success': False, 'error': '请先上传 Excel 文件'})
                return

            # 1. 复制data.json到部署目录
            import shutil
            shutil.copy(json_path, DEPLOY_DIR / 'data.json')

            # 2. 运行generate_articles_json.py生成articles.json
            gen_script = WORKSPACE / 'generate_articles_json.py'
            if gen_script.exists():
                result = subprocess.run(
                    [sys.executable, str(gen_script)],
                    capture_output=True, text=True, cwd=WORKSPACE
                )

            # 3. Git提交推送
            os.chdir(DEPLOY_DIR)
            git_files = ['data.json', 'articles.json']
            subprocess.run(['git', 'add'] + git_files, check=True, capture_output=True)
            subprocess.run(
                ['git', 'commit', '-m', f'update: 宣传看板数据更新 {datetime.now().strftime("%Y-%m-%d %H:%M")}'],
                check=True, capture_output=True
            )
            subprocess.run(['git', 'push', 'origin', 'main'], check=True, capture_output=True)

            self.send_json({
                'success': True,
                'message': '部署成功！(新架构: data.json + articles.json)',
                'url': 'https://kimiwayde-xu.github.io/csair-dashboard/'
            })
        except subprocess.CalledProcessError as e:
            self.send_json({
                'success': False,
                'error': f'Git 错误: {e.stderr.decode() if e.stderr else str(e)}'
            })
        except Exception as e:
            self.send_json({
                'success': False,
                'error': str(e)
            })


def parse_excel(file_path):
    """解析 Excel 文件，返回 JSON 数据结构"""
    if not HAS_OPENPYXL:
        raise Exception('需要安装 openpyxl: pip3 install openpyxl')

    wb = load_workbook(file_path, data_only=True)

    result = {
        "platform_distribution": {"platforms": [], "values": [], "percentages": []},
        "department_stats": {"departments": [], "values": []},
        "monthly_data": {"2025": {}, "2026": {}},
        "year_over_year": {},
        "quarter_yoy": {},
        "summary": {},
        "department_yoy": {}
    }

    # 解析数据源工作表
    ws = wb['数据源']

    # 平台分布 (行6-10, 列B-C)
    platforms = []
    platform_values = []
    for row in range(6, 11):
        name = ws.cell(row=row, column=2).value
        value = ws.cell(row=row, column=3).value
        if name:
            platforms.append(name)
            platform_values.append(value or 0)

    total_platform = sum(platform_values)
    platform_pct = [round(v/total_platform*100, 1) if total_platform else 0 for v in platform_values]

    result["platform_distribution"] = {
        "platforms": platforms,
        "values": platform_values,
        "percentages": platform_pct
    }

    # 全年科室上稿 (行6-17, 列E-F)
    dept_names = []
    dept_values = []
    for row in range(6, 18):
        name = ws.cell(row=row, column=5).value
        value = ws.cell(row=row, column=6).value
        if name:
            dept_names.append(name)  # 列E已是全称，直接使用
            dept_values.append(value or 0)

    result["department_stats"] = {
        "departments": dept_names,
        "values": dept_values
    }

    # 当月科室上稿 (行6-17, 列H-I)
    monthly_dept = []
    for row in range(6, 18):
        name = ws.cell(row=row, column=8).value
        value = ws.cell(row=row, column=9).value
        if name:
            monthly_dept.append(value or 0)

    # 官方同比数据（数据源工作表 行70-81）
    # 行70: 1月同比, 行71: 2月同比, ... 行74: 5月同比, ...
    ws_source = wb['数据源']
    yoy_monthly = {}
    monthly_yoy_totals = {}  # 各月总稿件同比
    
    # 从行42-53读取2025年各月总稿件量（O列）
    monthly_2025_totals = {}
    for row in range(42, 54):
        month = ws_source.cell(row=row, column=1).value
        if month and isinstance(month, str) and '月' in month:
            total = ws_source.cell(row=row, column=15).value or 0  # O列
            monthly_2025_totals[month] = int(total) if isinstance(total, (int, float)) else 0
    
    # 从行56-67读取2026年各月总稿件量（O列）
    monthly_2026_totals = {}
    for row in range(56, 68):
        month = ws_source.cell(row=row, column=1).value
        if month and isinstance(month, str) and '月' in month:
            total = ws_source.cell(row=row, column=15).value or 0  # O列
            monthly_2026_totals[month] = int(total) if isinstance(total, (int, float)) else 0
    
    # 计算总稿件同比
    for month in range(1, 13):
        month_name = f"{month}月"
        total_2025 = monthly_2025_totals.get(month_name, 0)
        total_2026 = monthly_2026_totals.get(month_name, 0)
        monthly_yoy_totals[month_name] = total_2026 - total_2025
    
    # 各科室同比
    for row in range(70, 82):  # 行70到81（1月到12月同比）
        month = ws_source.cell(row=row, column=1).value
        if month and isinstance(month, str) and '月' in month:
            yoy_values = [ws_source.cell(row=row, column=c).value or 0 for c in range(2, 14)]  # 列2到13共12个科室
            yoy_monthly[month] = yoy_values
    
    # 更新 year_over_year 中的 departments 为官方数据
    for month_name, yoy_values in yoy_monthly.items():
        result['year_over_year'][month_name] = {
            'total': monthly_yoy_totals.get(month_name, 0),
            'departments': yoy_values
        }

    # 科室同比（当月）- 使用官方数据
    result['department_yoy'] = yoy_monthly  # 所有月份的同比数据
    result['monthly_yoy_totals'] = monthly_yoy_totals  # 各月总稿件同比（供右上角模块使用）

    # 当月科室同比 (行6-17, 列K-L) - 保留作为备用
    yoy_dept = []
    for row in range(6, 18):
        name = ws.cell(row=row, column=11).value
        value = ws.cell(row=row, column=12).value
        if name:
            yoy_dept.append(value or 0)

    # 汇总数据
    total_articles = ws.cell(row=5, column=15).value or 0  # O5
    current_month_articles = ws.cell(row=6, column=15).value or 0  # O6
    monthly_avg = ws.cell(row=7, column=15).value or 0  # O7

    result["summary"] = {
        "total_articles": int(total_articles) if isinstance(total_articles, (int, float)) else 0,
        "monthly_avg": round(float(monthly_avg), 2) if isinstance(monthly_avg, (int, float)) else 0,
        "current_month_articles": int(current_month_articles) if isinstance(current_month_articles, (int, float)) else 0,
        "current_year": 2026,
        "current_month": datetime.now().month
    }

    # 平台占比 (行12, 列N-R)
    platform_pct_row = [ws.cell(row=12, column=c).value for c in range(14, 19)]
    if platform_pct_row[0]:
        result["platform_distribution"]["percentages"] = [round(p*100, 1) for p in platform_pct_row if p]

    # 当月平台明细 (行24-29)
    # 当前月份平台分布已废弃(current_month_platform)，前端用platform_distribution
    # 注：如有需要可以从月度获取
    # current_month_platform相关代码于2026-06-08移除

    # 2025年月度数据 (行42-53) - 使用O列作为total（官方总稿件量）
    for row in range(42, 54):
        month = ws.cell(row=row, column=1).value
        if month and isinstance(month, str) and '月' in month:
            month_name = month
            # O列(列15)是官方总稿件量
            total = ws.cell(row=row, column=15).value or 0
            # 列2到13共12个科室（列14是表头"X月稿件量"）
            depts = [ws.cell(row=row, column=c).value or 0 for c in range(2, 14)]
            result["monthly_data"]["2025"][month_name] = {
                "total": int(total) if isinstance(total, (int, float)) else 0,
                "departments": depts
            }

    # 2026年月度数据 - 从数据源工作表读取（行56-67，O列）
    # 不再从2026汇总表统计，使用官方预设数据
    for row in range(56, 68):
        month = ws.cell(row=row, column=1).value
        if month and isinstance(month, str) and '月' in month:
            month_name = month
            # O列(列15)是官方总稿件量
            total = ws.cell(row=row, column=15).value or 0
            depts = [ws.cell(row=row, column=c).value or 0 for c in range(2, 14)]  # 列2到13共12个科室
            result["monthly_data"]["2026"][month_name] = {
                "total": int(total) if isinstance(total, (int, float)) else 0,
                "departments": depts
            }

    # 同比数据 - 完全使用官方数据（数据源工作表）
    # 不再实时计算，直接从Excel读取
    # result["year_over_year"] 将在后面通过官方数据填充

    # 季度同比
    quarters = ["一季度", "二季度", "三季度", "四季度"]
    for i, q in enumerate(quarters):
        months_in_q = [f"{m}月" for m in range(i*3+1, i*3+4)]
        total_2025 = sum(result["monthly_data"]["2025"].get(m, {}).get("total", 0) for m in months_in_q)
        total_2026 = sum(result["monthly_data"]["2026"].get(m, {}).get("total", 0) for m in months_in_q)
        result["quarter_yoy"][q] = {
            "2025": total_2025,
            "2026": total_2026,
            "yoy": total_2026 - total_2025
        }

    # 从2026汇总表统计各月各平台稿件数
    monthly_platform_2026 = {m: {} for m in range(1, 13)}
    ws_2026 = wb['2026汇总表']
    
    for row in range(4, ws_2026.max_row + 1):
        date_val = ws_2026.cell(row=row, column=5).value
        platform = ws_2026.cell(row=row, column=3).value  # 媒体名称
        
        if date_val and platform:
            try:
                if isinstance(date_val, int):
                    # datetime already imported at module level
                    from datetime import timedelta
                    dt = datetime(1899, 12, 30) + timedelta(days=date_val)
                else:
                    dt = date_val
                month = dt.month
                monthly_platform_2026[month][platform] = monthly_platform_2026[month].get(platform, 0) + 1
            except:
                pass
    
    # 添加各月平台数据到 monthly_data
    for month in range(1, 13):
        month_name = f"{month}月"
        if month_name in result["monthly_data"]["2026"]:
            result["monthly_data"]["2026"][month_name]["platform"] = monthly_platform_2026[month]
    
    # 当前月份平台数据
    # 当前月份平台分布已废弃，前端不使用
    
    # 存储所有月份的平台数据供前端切换
    result["monthly_platform"] = monthly_platform_2026

    wb.close()
    return result


def get_html_template():
    return '''<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>宣传看板数据更新</title>
    <style>
        * { margin: 0; padding: 0; box-sizing: border-box; }
        body {
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif;
            background: linear-gradient(135deg, #0f172a 0%, #1e293b 100%);
            min-height: 100vh;
            color: #e2e8f0;
            padding: 40px 20px;
        }
        .container { max-width: 700px; margin: 0 auto; }
        h1 { font-size: 28px; margin-bottom: 8px; }
        .subtitle { color: #64748b; margin-bottom: 32px; }
        .card {
            background: rgba(30, 41, 59, 0.8);
            border: 1px solid rgba(148, 163, 184, 0.2);
            border-radius: 12px;
            padding: 24px;
            margin-bottom: 20px;
        }
        .card-title {
            font-size: 13px;
            color: #94a3b8;
            text-transform: uppercase;
            letter-spacing: 1px;
            margin-bottom: 16px;
        }
        .drop-zone {
            border: 2px dashed rgba(148, 163, 184, 0.3);
            border-radius: 12px;
            padding: 48px 24px;
            text-align: center;
            transition: all 0.3s;
            cursor: pointer;
        }
        .drop-zone:hover, .drop-zone.dragover {
            border-color: #3b82f6;
            background: rgba(59, 130, 246, 0.05);
        }
        .drop-zone-icon { font-size: 48px; margin-bottom: 12px; }
        .drop-zone-text { color: #94a3b8; }
        .drop-zone-hint { font-size: 12px; color: #475569; margin-top: 8px; }
        #fileInput { display: none; }
        .btn {
            background: linear-gradient(135deg, #3b82f6, #2563eb);
            border: none;
            color: white;
            padding: 14px 28px;
            font-size: 16px;
            border-radius: 8px;
            cursor: pointer;
            font-weight: 500;
            transition: all 0.3s;
            width: 100%;
        }
        .btn:hover:not(:disabled) {
            transform: translateY(-2px);
            box-shadow: 0 8px 25px rgba(59, 130, 246, 0.4);
        }
        .btn:disabled { opacity: 0.5; cursor: not-allowed; }
        .btn.success { background: linear-gradient(135deg, #22c55e, #16a34a); }
        .status {
            padding: 16px;
            border-radius: 8px;
            margin-top: 16px;
            font-size: 14px;
            display: none;
        }
        .status.show { display: block; }
        .status.success { background: rgba(34, 197, 94, 0.15); border: 1px solid rgba(34, 197, 94, 0.3); color: #22c55e; }
        .status.error { background: rgba(239, 68, 68, 0.15); border: 1px solid rgba(239, 68, 68, 0.3); color: #ef4444; }
        .status.info { background: rgba(59, 130, 246, 0.15); border: 1px solid rgba(59, 130, 246, 0.3); color: #60a5fa; }
        .preview { display: grid; grid-template-columns: repeat(4, 1fr); gap: 12px; margin-top: 16px; }
        .preview-item {
            background: rgba(15, 23, 42, 0.5);
            padding: 16px;
            border-radius: 8px;
            text-align: center;
        }
        .preview-value { font-size: 28px; font-weight: 700; color: #3b82f6; }
        .preview-label { font-size: 11px; color: #64748b; margin-top: 4px; }
        .link-box {
            background: rgba(34, 197, 94, 0.1);
            border: 1px solid rgba(34, 197, 94, 0.3);
            border-radius: 8px;
            padding: 20px;
            margin-top: 16px;
            text-align: center;
            display: none;
        }
        .link-box.show { display: block; }
        .link-box a { color: #22c55e; font-size: 18px; font-weight: 600; }
        .link-box-hint { font-size: 12px; color: #64748b; margin-top: 8px; }
        .spinner {
            display: inline-block;
            width: 16px;
            height: 16px;
            border: 2px solid rgba(255,255,255,0.3);
            border-top-color: white;
            border-radius: 50%;
            animation: spin 0.8s linear infinite;
            margin-right: 8px;
            vertical-align: middle;
        }
        @keyframes spin { to { transform: rotate(360deg); } }
    </style>
</head>
<body>
    <div class="container">
        <h1>📊 宣传看板数据更新</h1>
        <p class="subtitle">拖入 Excel 文件，一键更新线上看板</p>

        <div class="card">
            <div class="card-title">Step 1 / 上传 Excel</div>
            <div class="drop-zone" id="dropZone">
                <div class="drop-zone-icon">📁</div>
                <div class="drop-zone-text">拖入或点击选择 Excel 文件</div>
                <div class="drop-zone-hint">支持 .xlsx 格式，需包含「数据源」工作表</div>
            </div>
            <input type="file" id="fileInput" accept=".xlsx">
            <div id="statusUpload" class="status"></div>
            <div class="preview" id="preview" style="display:none;">
                <div class="preview-item">
                    <div class="preview-value" id="statTotal">-</div>
                    <div class="preview-label">总稿件量</div>
                </div>
                <div class="preview-item">
                    <div class="preview-value" id="statAvg">-</div>
                    <div class="preview-label">月均</div>
                </div>
                <div class="preview-item">
                    <div class="preview-value" id="statMonth">-</div>
                    <div class="preview-label">当月</div>
                </div>
                <div class="preview-item">
                    <div class="preview-value" id="statYoy">-</div>
                    <div class="preview-label">同比</div>
                </div>
            </div>
        </div>

        <div class="card">
            <div class="card-title">Step 2 / 部署到线上</div>
            <button class="btn" id="deployBtn" disabled onclick="deploy()">
                🚀 部署更新
            </button>
            <div id="statusDeploy" class="status"></div>
            <div class="link-box" id="linkBox">
                <a href="https://kimiwayde-xu.github.io/csair-dashboard/" target="_blank">
                    https://kimiwayde-xu.github.io/csair-dashboard/
                </a>
                <div class="link-box-hint">点击访问更新后的看板（约 30 秒生效）</div>
            </div>
        </div>
    </div>

    <script>
        const dropZone = document.getElementById('dropZone');
        const fileInput = document.getElementById('fileInput');
        let uploaded = false;

        dropZone.onclick = () => fileInput.click();
        dropZone.ondragover = (e) => { e.preventDefault(); dropZone.classList.add('dragover'); };
        dropZone.ondragleave = () => dropZone.classList.remove('dragover');
        dropZone.ondrop = (e) => {
            e.preventDefault();
            dropZone.classList.remove('dragover');
            if (e.dataTransfer.files[0]) uploadFile(e.dataTransfer.files[0]);
        };
        fileInput.onchange = (e) => { if (e.target.files[0]) uploadFile(e.target.files[0]); };

        function showStatus(el, type, msg) {
            const s = document.getElementById(el);
            s.className = 'status show ' + type;
            s.textContent = msg;
        }

        async function uploadFile(file) {
            const formData = new FormData();
            formData.append('file', file);

            showStatus('statusUpload', 'info', '⏳ 解析中...');

            try {
                const res = await fetch('/upload', { method: 'POST', body: formData });
                const data = await res.json();

                if (data.success) {
                    showStatus('statusUpload', 'success', '✅ ' + data.message);
                    document.getElementById('preview').style.display = 'grid';
                    document.getElementById('statTotal').textContent = data.summary.total_articles || '-';
                    document.getElementById('statAvg').textContent = data.summary.monthly_avg || '-';
                    document.getElementById('statMonth').textContent = data.summary.current_month_articles || '-';
                    document.getElementById('statYoy').textContent = (data.summary.yoy || 0) + '%';
                    document.getElementById('deployBtn').disabled = false;
                    uploaded = true;
                } else {
                    showStatus('statusUpload', 'error', '❌ ' + data.error);
                }
            } catch (err) {
                showStatus('statusUpload', 'error', '❌ 上传失败: ' + err.message);
            }
        }

        async function deploy() {
            if (!uploaded) return;

            const btn = document.getElementById('deployBtn');
            btn.disabled = true;
            btn.innerHTML = '<span class="spinner"></span>部署中...';
            showStatus('statusDeploy', 'info', '⏳ 正在推送到 GitHub...');
            document.getElementById('linkBox').classList.remove('show');

            try {
                const res = await fetch('/deploy', { method: 'POST' });
                const data = await res.json();

                if (data.success) {
                    showStatus('statusDeploy', 'success', '✅ ' + data.message);
                    document.getElementById('linkBox').classList.add('show');
                    btn.textContent = '✅ 已部署';
                    btn.classList.add('success');
                } else {
                    showStatus('statusDeploy', 'error', '❌ ' + data.error);
                    btn.disabled = false;
                    btn.textContent = '🚀 部署更新';
                }
            } catch (err) {
                showStatus('statusDeploy', 'error', '❌ 部署失败: ' + err.message);
                btn.disabled = false;
                btn.textContent = '🚀 部署更新';
            }
        }
    </script>
</body>
</html>'''


def main():
    os.chdir(WORKSPACE)

    print(f"""
╔══════════════════════════════════════════════════════╗
║     📊 宣传看板数据更新服务                            ║
╠══════════════════════════════════════════════════════╣
║  访问: http://127.0.0.1:{PORT}                         ║
║  按 Ctrl+C 停止服务                                   ║
╚══════════════════════════════════════════════════════╝
""")

    server = HTTPServer(('127.0.0.1', PORT), UploadHandler)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n服务已停止")
        server.shutdown()


if __name__ == '__main__':
    main()
