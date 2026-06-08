#!/usr/bin/env python3
"""
从Excel生成articles.json(科室稿件详情)
用于点击科室柱状图时显示该科室所有稿件

数据规则:
1. 从第4行开始解析,智能检测数据边界
2. 过滤统计行(第一列包含月份关键词)
3. 连续5行空行时停止解析
4. 无日期+无平台 → 标注"协助宣传工作"
5. 显示字段:科室、日期、平台、合计

改进:
- 移除硬编码行数限制(方案C)
- 智能适应新增数据行
- 统计行位置变化不影响解析
"""

import openpyxl
import json
from datetime import datetime, timedelta
from pathlib import Path

# ============================================
# 科室列定义（列6-17，对应2026汇总表第2行表头）
# 动态读取，保证与Excel保持一致
# ============================================
DEPT_COLUMNS = list(range(6, 18))  # 列6到17

# 月份关键词(用于识别统计行)

# 月份关键词(用于识别统计行)
MONTH_KEYWORDS = [
    '1月', '2月', '3月', '4月', '5月', '6月',
    '7月', '8月', '9月', '10月', '11月', '12月',
    '一月', '二月', '三月', '四月', '五月', '六月',
    '七月', '八月', '九月', '十月', '十一月', '十二月'
]

def convert_date(date_value):
    """日期转换函数(处理多种格式)"""
    if date_value is None:
        return None

    # datetime对象
    if isinstance(date_value, datetime):
        return date_value.strftime('%Y-%m-%d')

    # Excel序列号
    if isinstance(date_value, (int, float)):
        try:
            base_date = datetime(1899, 12, 30)
            return (base_date + timedelta(days=int(date_value))).strftime('%Y-%m-%d')
        except:
            return None

    # 字符串
    if isinstance(date_value, str) and date_value.strip():
        return date_value.strip()

    return None

def is_stat_row(ws, row_idx):
    """判断是否为统计行(第一列包含月份关键词)"""
    cell_value = ws.cell(row=row_idx, column=1).value
    if cell_value:
        cell_str = str(cell_value)
        for keyword in MONTH_KEYWORDS:
            if keyword in cell_str:
                return True
    return False

def read_dept_names(ws):
    """从Excel表头行（第2行，列6-17）读取科室名称"""
    dept_names = {}
    for col in DEPT_COLUMNS:
        name = ws.cell(row=2, column=col).value
        if name:
            dept_names[col] = str(name).strip()
        else:
            print(f"⚠️  列{col} 表头为空，使用默认名称 科室{col-5}")
            dept_names[col] = f"科室{col-5}"
    return dept_names


def generate_articles_json(excel_path, output_path):
    """生成articles.json"""

    # 加载Excel
    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    ws = wb['2026汇总表']

    # 从Excel表头动态读取科室名称（列6-17，第2行）
    EXCEL_COL_TO_DEPT = read_dept_names(ws)

    print(f"📋 从Excel读取的科室名称:")
    for col in DEPT_COLUMNS:
        print(f"  列{col}: {EXCEL_COL_TO_DEPT[col]}")
    print()

    # 初始化articles字典(使用Excel名称)
    articles = {dept_name: [] for dept_name in EXCEL_COL_TO_DEPT.values()}

    # 统计
    stats = {
        'total_rows': 0,
        'valid_rows': 0,
        'stat_rows': 0,
        'skipped_rows': 0,
        'assist_count': 0,  # 协助宣传工作数量
        'empty_consecutive': 0  # 连续空行计数
    }

    # 解析数据(从第4行开始,智能检测边界)
    # 不再硬编码行数上限,改为连续空行检测
    for row_idx in range(4, ws.max_row + 1):
        stats['total_rows'] += 1

        # 跳过统计行
        if is_stat_row(ws, row_idx):
            stats['stat_rows'] += 1
            continue

        # 读取基础信息
        title = ws.cell(row=row_idx, column=2).value
        platform = ws.cell(row=row_idx, column=3).value
        date_value = ws.cell(row=row_idx, column=5).value
        link = ws.cell(row=row_idx, column=18).value  # R列：新闻链接

        # 跳过空行
        if not title:
            stats['skipped_rows'] += 1
            stats['empty_consecutive'] += 1
            # 连续50行空行,停止解析(提高阈值以适应不同格式)
            if stats['empty_consecutive'] >= 50:
                print(f"  检测到连续50行空行,停止解析(行{row_idx})")
                break
            continue
        else:
            # 有标题,重置空行计数
            stats['empty_consecutive'] = 0

        stats['valid_rows'] += 1

        # 转换日期
        date_str = convert_date(date_value)

        # 判断是否为“协助宣传工作”（无日期+无平台）
        is_assist = (not date_str or date_str == '未标注') and (not platform or platform == '未标注')

        if is_assist:
            stats['assist_count'] += 1

        # 检查每个科室列
        for col in DEPT_COLUMNS:
            author = ws.cell(row=row_idx, column=col).value
            dept_name = EXCEL_COL_TO_DEPT[col]

            if author:  # 有姓名 = 该科室有此稿件

                article_data = {
                    'title': str(title)[:100] if title else '标题未知',
                    'platform': str(platform) if platform and str(platform).strip() else '协助宣传',
                    'date': date_str if date_str else '未标注'
                }

                # 添加链接（如果有）
                if link and str(link).strip():
                    article_data['link'] = str(link).strip()

                # 如果是协助宣传工作,添加标注
                if is_assist:
                    article_data['is_assist'] = True

                articles[dept_name].append(article_data)

    wb.close()

    # 保存JSON
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(articles, f, ensure_ascii=False, indent=2)

    # 输出统计
    print(f"✅ 生成 articles.json 完成")
    print(f"  总行数: {stats['total_rows']}")
    print(f"  有效行数: {stats['valid_rows']}")
    print(f"  统计行数: {stats['stat_rows']}")
    print(f"  跳过行数: {stats['skipped_rows']}")
    print(f"  协助宣传工作: {stats['assist_count']} 条")

    total_articles = sum(len(v) for v in articles.values())
    print(f"  总稿件数: {total_articles}")

    # 输出各科室稿件数量
    print(f"\n📊 各科室稿件数量:")
    for dept, article_list in sorted(articles.items(), key=lambda x: len(x[1]), reverse=True):
        assist_count = sum(1 for a in article_list if a.get('is_assist'))
        if assist_count > 0:
            print(f"  {dept}: {len(article_list)} 篇(含 {assist_count} 条协助宣传工作)")
        else:
            print(f"  {dept}: {len(article_list)} 篇")

    return articles

if __name__ == '__main__':
    excel_path = '/Volumes/way的固态/工作/2026/宣传/2026部门宣传数据可视化看板.xlsx'
    output_path = Path(__file__).parent / 'vercel-deploy' / 'articles.json'

    generate_articles_json(excel_path, output_path)
